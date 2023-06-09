from googlemap import get_location
from googlemap import search_places
import pandas as pd
from openpyxl import load_workbook

def mkdislike():
    df = pd.read_excel("data/imput.xlsx")

    api_key = ""
    address = df.iloc[0,2]
    radius = df.iloc[1,2]
    keywords = df.iloc[3,2].split(",")
    location = get_location(api_key, address)
    places = search_places(api_key, keywords, location, radius)
    wb = load_workbook('data/result.xlsx')
    ws = wb['Sheet1']

    # セルの開始位置
    start_row = 7
    start_column = 5

    # 値の書き込み
    for i, d in enumerate(places):
        row = start_row + i
        for j, key in enumerate(d.keys()):
            column = start_column + j
            cell = ws.cell(row=row, column=column)
            cell.value = d[key]

    # Excelファイルの保存
    wb.save('data/result.xlsx')